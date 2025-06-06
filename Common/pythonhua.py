from Common.do_path import conf_path,data_path
from Common.do_config import conf

import base64
import os
import hashlib
import time
import requests
import random
import threading
from contextlib import closing
from openpyxl import load_workbook, Workbook
import cv2
import shutil
import yaml
import ffmpeg
import pymysql




class myERROR(Exception):
    pass

# yaml文件操作
class MyYaml:
    def __init__(self, path, name, langue='utf-8', flag=False):
        self.path = path
        self.name = name
        self.paths = os.path.join(path, name)
        self.langue = langue

        # print(self.paths)

        # 判断文件名称是否存在文件目录下，不存在则新建
        plist = os.listdir(path)
        if str(name) not in plist:
            with open(self.paths, mode='w', encoding=self.langue) as f:
                f.write('')

    def readYaml(self):
        with open(self.paths, mode='r', encoding=self.langue) as f:
            conf = yaml.load(f, Loader=yaml.FullLoader)
        return (conf)

    def writeYaml(self, datas, mode='w'):
        with open(self.paths, mode=mode, encoding=self.langue) as f:
            conf = yaml.dump(datas, allow_unicode=True)
            f.write(conf)
yamlData = MyYaml(conf_path, "myConfig.yml")

class YamlPath:
    def __init__(self,path,name):
        self.yaml_data = MyYaml(path, name)
        self.yaml_read_data = self.yaml_data.readYaml()['yamlData']

        self.yaml_read_sql_data = self.yaml_read_data["sqlData"]
        self.yaml_read_sql_conns = self.yaml_read_sql_data["conns"]
        self.yaml_read_sql_database_makdir = self.yaml_read_sql_data["databaseMakdir"]
        self.yaml_read_sql_table_makdir = self.yaml_read_sql_data["tableMakdir"]
yaml_run = YamlPath(conf_path,"myConfig.yml")


class DoExcel:
    def __init__(self, paths, sheet, flag=False):
        """
        excel处理类
        :param paths: flag：False  excel完成的工作路径     flag：存放在TestDatas目录下True excel的文件名
        :param sheet: 工作表名称
        :param flag: 开关  主要控制paths路径
        """

        if flag:
            paths = os.path.join(data_path, paths)
        self.paths = paths
        self.sheet = sheet
        self.wb = load_workbook(self.paths)
        self.sheet = self.wb[self.sheet]

    def read_excel(self, row, column):
        """
        read_excel:读取单元格的值
        row:单元格行数
        column:单元格列
        """

        values = self.sheet.cell(row, column).value
        return values

    def write_excel(self, row, column, values):
        """
        write_excel:修改单元格的值
        row:单元格行数
        column:单元格列
        values:需要修改的值
        """

        self.sheet.cell(row, column).value = values
        self.wb.save(self.paths)

    def title_write(self, titleNames, rows=1):
        """
        新建Excel标签行
        :param titleNames: 标题列名
        :param rows: 标签行
        :return:
        """
        if isinstance(titleNames, list):
            titles = titleNames
        else:
            titles = []
            titles.append(titleNames)
        for i in range(len(titleNames)):
            self.write_excel(rows, i + 1, titles[i])

        self.wb.save(self.paths)

    def all_excel(self):  # 推荐使用
        # 定义一个空列表,获取title标签名
        lists = []
        for item in list(self.sheet.rows)[0]:
            lists.append(item.value)

        # 放已经转换好的测试数据
        all_list = []
        for item in list(self.sheet.rows)[1:]:
            # 定义一个空列表放测试数据
            list_data = []
            for i in item:
                list_data.append(i.value)
            # dict(zip(value1,value2))  zip将value1和valu2组合，dict将zip组合的数据以字典的形式呈现
            res = dict(zip(lists, list_data))

            # json.loads 将字符串转为json格式
            # res["request_data"] = json.loads(res["request_data"])
            # res["expected"] = json.loads(res["expected"])
            # json.dumps 将json格式转为字符串
            # res["request_data"] = json.dumps(res["request_data"])

            all_list.append(res)
        return (all_list)

    def all_excel_one(self):
        # 定义一个空列表,获取title标签名
        lists = []
        for item in list(self.sheet.rows)[0]:
            lists.append(item.value)

        # 定义空列表放所有字典数据
        all_list = []
        # 取除了标题行的所有行的数据
        for item in list(self.sheet.rows)[1:]:
            dicts = {}
            # 对每一行进行循环得到每个单元格的值
            for index in range(len(item)):
                dicts[lists[index]] = item[index].value
            # excel读取的都是字符串或者数据,利用eval去掉字符串得到原有格式
            # dicts["msg"] = eval(dicts["msg"])
            # dicts["request_data"] = json.loads(dicts["request_data"])
            # dicts["expected"] = json.loads(dicts["expected"])
            all_list.append(dicts)

        return (all_list)

    def all_excel_two(self):
        # 最大行
        row = self.sheet.max_row

        # 最大列
        column = self.sheet.max_column
        lists = []
        for item in range(column):
            lists.append(self.sheet.cell(1, item + 1).value)
        all_lists = []
        for item in range(2, row + 1):
            dicts = {}
            for i in range(1, column + 1):
                dicts[lists[i - 1]] = self.sheet.cell(item, i).value
            # dicts["msg"] = eval(dicts["msg"])
            # dicts["request_data"] = json.loads(dicts["request_data"])
            # dicts["expected"] = json.loads(dicts["expected"])
            all_lists.append(dicts)

        return (all_lists)

    def max_row(self):
        # 最大行
        row = self.sheet.max_row
        return row

    def max_column(self):
        # 最大列
        column = self.sheet.max_column
        return column

    def clos_excle(self):
        self.wb.close()

    def sava_data(self):
        self.wb.save(self.paths)

lock = threading.Lock()
class MyOs:

    def multi_thread(self, obj, data, num, flag=False):
        """
        并发的封装
        :param obj: 并发执行的函数
        :param data: 并非需要的数据
        :param num: 并发数
        :param flag: 用于执行obj函数传两个参数使用，默认一个参数
        :return: 执行
        """
        threads_list = []
        # flag :True 允许传两个参数
        if flag:
            for value in data:
                threads_list.append(threading.Thread(target=obj, args=(value, flag), name=f"进程{int(time.time())}"))
        else:
            for value in data:
                threads_list.append(threading.Thread(target=obj, args=(value,), name=f"进程{int(time.time())}"))
        # 开始线程计数
        num_one = 0
        # 结束线程计数
        num_two = 0
        # 每次执行的并发数
        num = int(num)
        # 开始线程计数，小于数据条数，就启动线程处理数据
        while num_one < len(data):
            # 每次启动线程的并发数
            for i in range(num):
                print("{}{}".format(threading.current_thread().name, num_one))
                threads_list[num_one].start()
                num_one += 1
                # 开始线程计数，大于数据条数，跳出循环
                if num_one >= len(data):
                    break
            # 每次等待线程的并发数结束
            for i in range(num):
                threads_list[num_two].join()
                num_two += 1
                # 开始线程计数，大于数据条数，跳出循环
                if num_two >= len(data):
                    break

    def filelist(self, file_dir):
        """
        查找文件夹下的所有文件，包含子目录、孙子目录下所有的文件
        :param file_dir: 文件夹路径
        :return: 返回文件列表
        """
        if os.path.exists(file_dir) == False:
            # 判断文件路径是否存在，不存在直接返回
            return "{}文件路径".format(file_dir)
        # 拉出根目录下的所有文件
        path_list = []
        for root, dirs, files in os.walk(file_dir):
            # root 根目录
            # dirs 子目录
            # files 子目录文件
            # 主要拉出根路径的文件
            if root == file_dir and dirs != []:
                # print(root)
                for i in range(len(files)):
                    paths = os.path.join(root, files[i])
                    path_list.append(paths)
                    # print(paths)
                # print("\n")
            # 主要拉出根目录下的所有子目录的文件
            if dirs == []:
                # print('目录路径    ：',root)
                # print('子目录名称  ：',dirs)
                # print('目录下的文件：',files)
                # print("\n")
                # print(root)
                for i in range(len(files)):
                    paths = os.path.join(root, files[i])
                    path_list.append(paths)
                    # print(paths)
                # print("\n")
        return path_list

    def findStype(self, file_dir, stype="txt"):
        """
        查找文件后缀为指定格式文件
        :param file_dir: 文件路径
        :param stype: 文件格式 默认 txt
        :return: 输出符合条件的路径列表
        """
        if isinstance(file_dir, list):
            # 判断是否为列表
            file_dirs = file_dir
        elif isinstance(file_dir, str) and os.path.exists(file_dir):
            # 判断是否为字符串，并且文件路径是否存在
            file_dirs = []
            file_dirs.append(file_dir)
        else:
            print("输入的文件路径{}不对".format(file_dir))
            return []

        values = []
        for files in file_dirs:
            if os.path.exists(files):
                filesList = files.split(".")
                if filesList[len(filesList) - 1] == stype:
                    values.append(files)
        return values

    def outfile(self, fileFront):
        """
        随机更改文件，文件，字符
        :param moveFront: 需要更改的文件、文件夹路径
        :return:
        """
        outstr = ''
        if os.path.isfile(fileFront):
            # 文件随机命名
            value = fileFront.split('.')
            outstr = "{}{}_副本{}{}{}".format(outstr, value[0], random.randint(1, 9), random.randint(0, 9),
                                            random.randint(0, 9))
            for i in range(1, len(value) - 1):
                outstr = outstr + '.' + value[i]
            outstr = "{}.{}".format(outstr, value[len(value) - 1])

        elif os.path.isdir(fileFront):
            # 文件夹随机命名
            value = fileFront.split('/')
            for i in range(0, len(value) - 1):
                outstr = outstr + value[i] + '/'
            outstr = "{}{}_副本{}{}{}".format(outstr, value[len(value) - 1], random.randint(1, 9), random.randint(0, 9),
                                            random.randint(0, 9))

        else:
            # 字符随机
            outstr = "{}_副本{}{}{}".format(fileFront, random.randint(1, 9), random.randint(0, 9), random.randint(0, 9))

        return outstr

    def movefile(self, moveFront, moveAfter):
        """
        移动文件、文件夹
        :param moveFront: 需要移动的文件、文件夹路径
        :param moveAfter: 移动后的文件、文件夹路径
        :return:
        """
        try:
            shutil.move(moveFront, moveAfter)
        except:
            print("{} 在文件 {} 已存在".format(moveFront, moveAfter))
            return moveFront

    def copyfile(self, copyFront, copyAfter, overlay=True):
        """
        复制文件、文件夹
        :param copyFront: 需要复制的文件、文件夹路径
        :param copyAfter: 复制后的文件、文件夹路径
        :param overlay: True 复制    False 覆盖复制
        :return: 无
        """

        if overlay:
            shutil.copy(copyFront, copyAfter)
        else:
            shutil.copy2(copyFront, copyAfter)

    def renamefile(self, nameFront,flag = None, nameAfter=None):
        """
        文件、文件夹重命名
        :param nameFront: 需要命名的文件、文件夹路径
        :param flag: True 根目录创建宽高  Flase 当前文件目录创建宽高
        :param nameAfter: 命名后的文件、文件夹路径
        :return: 执行
        """

        if nameAfter == None and os.path.isfile(nameFront):
            try:
                # 读取视频图片md5
                md5 = self.filemd5(nameFront)
                # 读取视频图片信息
                values = self.get_video_info(nameFront)
                # print(values)
                # 判断视频图片信息[]
                if values !=[]:
                    # 视频图片文件的目录
                    folder_path = values[0]["folder_path"]
                    width = values[0]["width"]
                    height = values[0]["height"]
                    size = values[0]["size"]
                    # 读取视频图片文件格式
                    file_format = values[0]["file_format"]
                    image_formats = ["bmp","jpeg","jpg","png","tif","gif","pcx","tga","exif","fpx","svg","psd","cdr","pcd","dxf","ufo","eps","ai","raw","WMF","webp","avif","apng"]
                    video_formats = ["mp4","mov","m4v","wmv","avi","flv"]
                    # 读取本地视频图片文件的格式
                    str_values = self.str_split(nameFront,'.')
                    # 如果读取文件后缀不为[]
                    if str_values != []:
                        Suffix = str_values[-1]
                    else:
                        Suffix = values[0]["file_format"]
                    if file_format in image_formats:
                        file_name = "image"
                    elif file_format in video_formats:
                        file_name = "video"
                    else:
                        file_name = "material"
                    # 拼接文件名称+后缀
                    name = "{}_{}_{}_{}.{}".format(file_name,width,height,md5,Suffix)
                    # 为None 当前文件目录下创建，否者给定的目录进行创建
                    if flag == None:
                        all_path = os.path.dirname(folder_path)
                    else:
                        all_path = flag
                    # 文件目录和文件宽高拼接
                    folder_paths = os.path.join(all_path,"{}_{}".format(width,height))
                    # print(folder_paths)
                    # 拼接的目录存在不进行创建，不存在进行创建
                    self.is_path(folder_paths)
                    # 文件目录+文件名称进行拼接
                    file_paths = os.path.join(folder_paths,name)
                    os.rename(nameFront,file_paths)

            except:
                pass
        else:
            # 文件夹命名
            if os.path.isdir(nameFront) and os.path.isdir(nameAfter):
                os.rename(nameFront, nameAfter)
            # 文件命名
            elif os.path.isfile(nameFront) and os.path.isfile(nameAfter):
                os.rename(nameFront, nameAfter)
            else:
                print("{} 和 {}不能进行重命名".format(nameFront, nameAfter))

    def readfile(self, file_path):
        """
        读取文件文件内容
        :param file_path: 文本文件路径
        :return: 返回文本文件内容
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                return line

        f.close()

    def writefile(self, file_path, txt, isw=False):
        """
        写文本文件
        :param file_path: 问你路径
        :param txt: 写入的文本
        :param isw: True 覆盖写入   False 追加写入
        :return: 无
        """
        if isw:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(txt)
        else:
            with open(file_path, 'a', encoding='utf-8') as f:
                f.write("\n" + txt)
        f.close()

    def is_path(self, path):
        """
        判断目录是否存在，不存在就进行新建
        :param path: 文件路径
        :return: 执行
        """

        if os.path.isdir(path) == False:
            os.makedirs(path)

    def chagemd5(self, file_path):
        # 改变文件的MD5
        # filename 文件路径
        myfile = open(file_path, 'a')
        myfile.write("####&&&&")
        myfile.close

    def filemd5(self, file_path, flag=False):
        """
        读取文件MD5
        :param filename: 路径
        :param flag: 开关 True 改变MD5  False 不改变MD5 只读取MD5
        :return:返回MD5
        """

        if flag:
            self.chagemd5(file_path)
        hasher = hashlib.md5()
        with open(file_path, "rb") as file:
            buf = file.read()
            while len(buf) > 0:
                hasher.update(buf)
                buf = file.read()
        # print(hasher.hexdigest())
        return hasher.hexdigest()

    def str_split(self, data, ide='\\'):
        """
        数据中找到idt，找到返回找到的数据，没有找到返回[]
        :param data:数据源
        :param ide:按照ide对数据源切割
        :return:切割后的列表
        """
        data = str(data)
        if data.find(ide) == -1:
            return []
        else:
            value = data.split(ide)
            return value

    def str_strip(self,data,direction = "all"):
        data = str(data)
        if direction == "left":
            value = data.lstrip()
            return value
        elif direction == "right":
            value = data.rstrip()
            return value
        else:
            value = data.strip()
            return value

    def name_rand(self, name="Resource", out_file_path=None, flag=False):
        """
        文件名称拼接随机值
        :param name: 文件名称
        :param out_file_path: True 进行文件路径拼接   Flase 不进行文件路径拼接
        :param flag: True 进行拼接  Flase 不进行拼接
        :return: 输出随机命名的文件或者路径
        """

        if flag:
            file_name = "{}_{}{}{}{}".format(name, int(time.time()), random.randint(1, 9), random.randint(1, 9),
                                             random.randint(1, 9), random.randint(1, 9))
        else:
            file_name = name
        if out_file_path != None and os.path.isdir(out_file_path):
            file_name = os.path.join(out_file_path, file_name)

        return file_name

    def get_video_dimensions(self, url):
        # 通过视频ulr解析视频宽高
        video_capture = cv2.VideoCapture(url)
        width = int(video_capture.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(video_capture.get(cv2.CAP_PROP_FRAME_HEIGHT))
        video_capture.release()
        return width, height

    def get_url_md5(self, url):
        filename = url.split('/')[-1]  # 提取文件名

        response = requests.get(url)  # 发送GET请求

        md5_obj = hashlib.md5()  # 创建MD5对象
        md5_obj.update(response.content)  # 更新MD5对象

        return md5_obj.hexdigest()  # 返回MD5值

    def get_video_info(self, path):
        try:
            # 解析视频文件
            # probe = ffmpeg.probe('./测试视频文件.mp4')
            probe = ffmpeg.probe(path)
            # print(probe)  # 获取视频多媒体文件的信息
            format = probe['format']
            # print(format)

            filepath = format['filename']  # 文件路径
            # logger.info("文件路径为：{}".format(filepath))
            filenames = os.path.basename(filepath)  # 获取文件路径中的文件名称，带后缀
            filename = filenames.split(".")[0]  # 获取文件路径中的文件名称，去掉后缀
            # filename = os.path.splitext(filepath)

            folder_path = str.split(filepath, filenames)[0]  # 利用str截取字符串获取文件夹的路径

            fileformat = os.path.splitext(filepath)[1]  # 获取文件格式
            fileformat = fileformat.split(".")[1]  # 去掉圆点

            # logger.info("文件名称为：{}".format(filename))
            # logger.info("文件格式为：{}".format(fileformat))

            bit_rate = format['bit_rate']
            # print("比特率(bps)为:{}".format(int(bit_rate)))  # 单位 bps（每秒字节数）
            kbps = int(bit_rate) / 1000
            kbps = round(kbps)
            # logger.info("比特率为(kbps):{}".format(kbps))  # 单位 kbps（每秒字节数）kbps

            fps_a = int(probe['streams'][0]['r_frame_rate'].split('/')[0])
            fps_b = int(probe['streams'][0]['r_frame_rate'].split('/')[1])
            fps = fps_a / fps_b  # 获取帧率
            fps = round(fps)
            # logger.info("视频帧率为：{}".format(fps))

            streams = probe['streams'][0]
            width = streams['width']
            coded_width = streams['coded_width']
            height = streams['height']
            coded_height = streams['coded_height']
            # logger.info("视频宽为(px)：{}".format(width))
            # logger.info("视频高为(px)：{}".format(height))

            duration = format['duration']
            duration = round(float(duration), 2)  # 时长（单位秒）
            # logger.info("时长为(s)：{}".format(duration))

            # 通过比特率X时长/8 计算文件大小，不是很准确
            # file_size = kbps * duration / 8
            # print("通过比特率X时长/8 计算文件大小:{}kb".format(file_size))  # 得到文件的大小是 KB
            # print("通过比特率X时长/8/1024 计算文件大小:{}M".format(file_size / 1024))  # 计算得到的数据

            size_b = int(format['size'])  # 获取文件大小（单位字节B）
            size_kb = float(format['size']) / 1024  # 获取文件大小（单位字节KB）
            size = float(format['size']) / 1024 / 1024  # 获取文件大小（单位字节M）
            size = round(size, 2)  # 文件大小保留两位小数
            size_kb = round(size_kb, 2)  # 文件大小保留两位小数
            # logger.info("文件大小为(b)：{}".format(size_b))
            # logger.info("文件大小为(kb)：{}".format(size_kb))
            # logger.info("文件大小为(m)：{}".format(size))

            title = {"folder_path": "文件夹路径", "file_path": "文件路径", "file_name": "文件名称", "file_format": "文件后缀",
                     "kbps": "比特率(kbps)", "fps": "视频帧率", "width": "视频宽(px)", "height": "视频高(px)", "duration": "时长(s)",
                     "size_b": "文件大小(b)", "size_kb": "文件大小(kb)", "size": "文件大小(m)"}
            # logger.info("输出所有字段为：\n{}".format(title))
            dict_list = {}
            # dict_list["title"] = title
            dict_list["folder_path"] = folder_path
            dict_list["file_path"] = filepath
            dict_list["file_name"] = filename
            dict_list["file_format"] = fileformat
            dict_list["kbps"] = kbps
            dict_list["fps"] = fps
            dict_list["width"] = width
            dict_list["height"] = height
            dict_list["duration"] = duration
            dict_list["size_b"] = size_b
            dict_list["size_kb"] = size_kb
            dict_list["size"] = size
            lists = []
            lists.append(dict_list)
        except:
            # logger.info("文件异常：{}".format(path))
            return None
        return (lists)

    def file_info(self,input_path,out_path):
        """
        提取文件信息进行返回
        :param input_path: 输入文件信息
        :return: 返回文件信息

         目前全部字段，前三个一定返回
        ["name","path","format","md5","width","height","duration","size_b","size_kb","size"]
        """
        lock.acquire()
        image_formats = ["bmp", "jpeg", "jpg", "png", "tif", "gif", "pcx", "tga", "exif", "fpx", "svg", "psd", "cdr",
                         "pcd", "dxf", "ufo", "eps", "ai", "raw", "WMF", "webp", "avif", "apng"]
        video_formats = ["mp4", "mov", "m4v", "wmv", "avi", "flv"]
        material_formats  = image_formats + video_formats



        # 存放返回数据
        data_list = []
        # print(material_formats)
        # 文件名
        file_name = self.str_split(input_path)[-1]
        new_name = self.str_split(file_name,'.')
        if new_name != []:
            txt = ''
            if len(new_name) > 2:
                # 减1为了去除后缀，可能文件名有多个.所以用for
                txt = new_name[0]
                for i in range(1,len(new_name)-1):
                    txt = "{}.{}".format(txt,new_name[i])
            else:
                txt = new_name[0]
            file_name = self.str_strip(txt)

        # 文件后缀
        file_format = self.str_split(input_path,'.')[-1]
        # 文件md5
        md5 = self.filemd5(input_path)

        data_list.append(file_name)
        data_list.append(input_path)
        data_list.append(file_format)
        data_list.append(md5)

        if file_format in material_formats:
            try:
                # 素材信息
                material_info = self.get_video_info(input_path)[0]
                # print(material_info)
                width = material_info["width"]
                height = material_info["height"]
                duration = material_info["duration"]
                size_b = material_info["size_b"]
                size_kb = material_info["size_kb"]
                size = material_info["size"]

                data_list.append(width)
                data_list.append(height)
                data_list.append(duration)
                data_list.append(size_b)
                data_list.append(size_kb)
                data_list.append(size)
            except:
                codes = str({"msg":"读取视频信息错误"})
                data_list.append(codes)
                pass



        if os.path.isfile(out_path):
            csv_path = os.path.dirname(out_path)
            csv_name = self.str_split(out_path)[-1]
            mycsv = MyCsv(csv_path,csv_name,"gbk")
            mycsv.dowriteList(data_list)

        lock.release()

class MyCsv:
    def __init__(self, path, name, langue='utf-8', flag=False):
        self.path = path
        self.name = name
        self.paths = os.path.join(path, name)
        self.langue = langue

        # 判断文件名称是否存在文件目录下，不存在则新建
        plist = os.listdir(path)
        if str(name) not in plist:
            self.title("a", "b")

    def readCsv(self, swith_to="#", default="None", mode='r'):
        with open(self.paths, mode=mode, encoding=self.langue) as f:
            txtlines = f.readlines()
        title = txtlines[0].strip().split(',')

        all_list = []
        title_num = len(title)

        for i in range(1, len(txtlines)):
            # strip 截取首尾空字符
            temp = txtlines[i].strip()
            # startswith查找什么开头
            if temp.startswith(swith_to):
                continue
            # 以,号为分隔符截取字段
            temp_list = temp.split(',')
            num = len(temp_list)
            dict_list = {}
            for j in range(title_num):
                if j < int(num):
                    dict_list[title[j]] = temp_list[j]
                else:
                    dict_list[title[j]] = default

            all_list.append(dict_list)

        return all_list

    def readAll(self, mode='r'):
        with open(self.paths, mode=mode, encoding=self.langue) as f:
            value = f.read()
            return value

    def dowrite(self, *args):
        txt = ''
        for i in range(len(args)):
            if i == 0:
                txt = "{}".format(args[i])
            else:
                txt = "{},{}".format(txt, args[i])

        with open(self.paths, mode='a', encoding=self.langue) as f:
            f.write('\n{}'.format(txt))

    def dowriteList(self, data):
        txt = ''
        for i in range(len(data)):
            if i == 0:
                txt = "{}".format(data[i])
            else:
                txt = "{},{}".format(txt, data[i])

        with open(self.paths, mode='a', encoding=self.langue) as f:
            f.write('\n{}'.format(txt))

    def title(self, *args):
        txt = ''
        for i in range(len(args)):
            # 查找字符重复数
            counts = txt.count(args[i])
            # if counts > 0:
            #     raise Exception("{}字段重复".format(args[i]))
            if i == 0:
                txt = "{}".format(args[i])
            else:
                txt = "{},{}".format(txt, args[i])
        with open(self.paths, mode='w', encoding=self.langue) as f:
            f.write('{}'.format(txt))




# 并发
class MultiConcurrence(MyOs):
    def multi_move_file(self, input_path, out_path, num=3):
        """
        文件、文件夹移动
        :param input_path: 移动文件、文件夹前路径
        :param out_path: 移动文件、文件夹后路径
        :param num: 并发数
        :return: 执行
        """
        values = self.filelist(input_path)
        self.multi_thread(self.movefile, values, num, out_path)

    def multi_copy_file(self, input_path, out_path, num=3):
        """
        文件、文件夹拷贝
        :param input_path: 拷贝文件、文件夹前路径
        :param out_path: 拷贝文件、文件夹后路径
        :param num: 并发数
        :return: 执行
        """
        values = self.filelist(input_path)
        self.multi_thread(self.copyfile, values, num, out_path)

    def multi_rename_file(self, input_path, out_path=None, num=3, flag = False):
        """
        文件、文件夹重命名
        :param input_path:命名前的文件、文件夹
        :param out_path:明确的命名文件、文件夹
        :param num: 并发数
        :param flag: 指定的文件夹下
        :return: 执行
        """
        values = self.filelist(input_path)
        if flag == False:
            self.multi_thread(self.renamefile, values, num)
        else:
            self.multi_thread(self.renamefile,values,num,input_path)

    def multi_get_path_write_csv(self,input_path,out_path = None,file_name = None, num=3,flag = False):

        if out_path == None:
            out_path = input_path
        if file_name == None:
            name = "StatisticalPath.csv"
        else:
            name = file_name
        file_path = os.path.join(out_path,name)
        title_flag = os.path.isfile(file_path)

        mycsv = MyCsv(out_path,name)
        if flag == False or title_flag == False:
            mycsv.title("name","path","format","md5","width","height","duration","size_b","size_kb","size_m")

        datas = self.filelist(input_path)
        self.multi_thread(self.file_info,datas,num,file_path)

# 数据库操作
# sqlData = yamlData.readYaml()["yamlData"]["sqlData"]
class MyDb:
    # 1.连接数据库
    # print(yaml_run.yaml_read_sql_data)
    def __init__(self,flag = False):
        cons = yaml_run.yaml_read_sql_conns
        if flag:
            switch = flag
        else:
            switch = cons['switch']
        self.switch = switch
        self.login = cons[switch]
        self.database = cons[switch]["database"]
        self.conn = pymysql.connect(
            host=self.login["host"],
            port=self.login["port"],
            user=self.login["user"],
            password=self.login["password"],
            database=self.database,
            charset="utf8",
            cursorclass=pymysql.cursors.DictCursor  # 数组转为字典
        )

        # 2.创建游标
        self.cur = self.conn.cursor()

    # 获取单条数据结果
    def get_one_db(self,sql):
        self.conn.commit()
        # 3.执行sql语句
        self.cur.execute(sql)
        # 4.执行sql语句,获取数据结果
        return self.cur.fetchone()

    # 获取所有数据结果
    def get_all_db(self,sql):
        self.conn.commit()
        # 3.执行sql语句
        self.cur.execute(sql)
        # 4.执行sql语句,获取数据结果
        return self.cur.fetchall()

    # 获取数据结果条数
    def get_count_db(self,sql):
        self.conn.commit()
        # 3.执行sql语句
        count = self.cur.execute(sql)
        # 4.执行sql语句,获取数据结果
        return count

    # 获取所有数据库名
    def database_name(self):
        """
        列举数据库下所有数据库
        :return: 列表中返回所有数据库
        """
        sql = "show databases"
        result = self.get_all_db(sql)
        databases = []
        for i in range(len(result)):
            databases.append(result[i]["Database"])
        return databases

    # 切换数据库
    def database_use(self,datbase_name):
        """
        数据库切换
        :param datbase_name: 数据库名
        :return: 无
        """
        self.conn.commit()
        sql = "use {}".format(datbase_name)
        # 3.执行sql语句
        count = self.cur.execute(sql)

    # 排除内部数据库
    def database_where(self,data = False,sql_list = ['information_schema', 'mysql', 'sys','performance_schema']):
        """
        排除sql数据库内部数据库
        :param data: 数据库列表
        :param sql_list: 排除的数据库
        :return: 返回排除后的数据库
        """
        new_data = []
        if data == False:
            datas = mydb.database_name()
        else:
            datas = data

        try:
            for value in datas:
                if value not in sql_list:
                    new_data.append(value)
        except:
            pass

        return new_data



    # 获取所有数据表名
    def tabale_name(self,datbaseName = False):
        """
        列举数据库中所有数据表
        :param flag: 不传：取登录数据库中数据库，   传值：切换当前数据库
        :return: 返回数据库中所有数据表
        """

        dates = self.database_name()
        # print(yaml_run.yaml_read_sql_data)
        if datbaseName == False:
            datbaseName = self.database
        if datbaseName not in dates:
            raise myERROR("数据库{}不在{}中".format(datbaseName, dates))

        self.database_use(datbaseName)
        sql = "show tables"
        result = self.get_all_db(sql)
        tables = []
        if len(result) == 0:
            return []
        for key,val in result[0].items():
            keys = key
        for i in range(len(result)):
            tables.append(result[i][keys])
        return tables

    # 获取表的列表名
    def tabale_desc(self,tableName=False,databaseName =False):
        """
        获取表的字段，相当于sql中desc
        :param flag: 传：取传入的表名，不传：取配置中的表名
        :return: 返回所有表字段
        """
        try:
            if tableName:
                tableName = tableName
                if databaseName:
                    tables = self.tabale_name(databaseName)
                else:
                    tables = self.tabale_name()
                if tableName not in tables:
                    raise myERROR("传入的表名不对，请重新查入查询的表名")
            else:
                tableName = self.login["table"]
        except:
            raise myERROR("传入的表名不对，请重新查入查询的表名")

        self.get_all_db("desc {}".format(tableName))
        sql = "desc {}".format(tableName)
        result = self.get_all_db(sql)
        return result

    # 解析表中有多少字段
    def tabale_title(self,datas, flag=False):
        """
        解析表中有多少字段
        例如：[{'Field': 'id', 'Type': 'int unsigned', 'Null': 'NO', 'Key': 'PRI', 'Default': None, 'Extra': 'auto_increment'}]
        :param datas: 数据源
        :param flag: 不传默认表中字段，   传了取对应字段值
        :return: 返回表中取得的字段
        """
        tablename = []
        if flag:
            try:
                for values in datas:
                    tablename.append(values[flag])
            except:
                raise myERROR("输入的入参值不对：{}".format(datas))
        else:
            try:
                for values in datas:
                    tablename.append(values["Field"])
            except:
                raise myERROR("输入的入参值不对：{}".format(datas))
        return tablename

    # 将列表拼接成字符串
    def key_str(self,data, Symbol=","):
        """
        将列表拼接成字符串
        例如输入["id","name","tel"]  输出 id,name,tel
        :param data: 列表
        :param Symbol: 查询表的字符
        :return: 输出字符
        """
        if isinstance(data, list):
            datas = data
        else:
            datas = []
            datas.append(data)

        stsdata = ''
        for i in range(len(datas)):
            stsdata += datas[i]
            if i != len(datas) - 1:
                stsdata += Symbol
        return stsdata


    # 将列表拼接成值
    def value_str(self,data, Symbol=","):
        """
        将列表拼接成值
        例如输入["id","name","tel"]  输出 id,name,tel
        :param data: 列表
        :param Symbol: 查询表的字符
        :return: 输出字符
        """
        if isinstance(data, list):
            values = ''
            for i in range(len(data)):
                # 判断sql查询条件是否为str，是需要加上""查询
                if isinstance(data[i], str):
                    if data[i] =='':
                        value = 'Null'
                    else:
                        value = "\"{}\"".format(data[i])
                else:
                    value = data[i]

                if i==0:
                    values = value
                else:
                    values = "{}{}{}".format(values, Symbol, value)

            return values
        else:
            raise myERROR("{}不为列表".format(data))

    # 查询数据表中所有数据
    def select_all(self,table_title=False, table_name=False, database_name=False):
        """
        查询数据表中的所有数据
        :param table_title: 表名字段列表[]
        :param table_name: 表名
        :param database_name: 数据库名
        :return:
        """
        if database_name == False:
            database_name = self.database
            self.get_all_db("use {}".format(database_name))
        if table_name == False:
            try:
                table_name = self.login["table"]
            except:
                raise myERROR("请输入查询的数据表名table_name")
        if table_title == False:
            table_title = "*"
        else:
            table_title = self.key_str(table_title)

        sql = "select {} from {}".format(table_title, table_name)
        result = self.get_all_db(sql)
        return result

    # 执行数据操作
    def run_db(self, sql):
        # 3.执行sql语句
        self.cur.execute(sql)
        # 4.保存数据库
        self.conn.commit()
        return self.select_all()

    # 整理sql查询语句
    def value_sql_str(self,datas):
        """
        整理sql输出值
        :param datas: {"name": "小明", "age": 18, "tel": 13567823931}
        :return: ('name,age,tel', '"小明",18,13567823931')
        """
        sql_key = []
        sql_value = []
        for key,value in datas.items():
            sql_key.append(key)
            sql_value.append(value)
        return self.key_str(sql_key),self.value_str(sql_value)


    # 拼接sql的where的条件
    def where_str(self,wheres, mod="and"):
        """
        将[{"field":"id","operator" :">","value":1}, {"field":"name","operator" :"=","value":"jake"}]转为id>1 or name="jake"
        :param wheres: 转换条件  [{"field":"id","operator" :">","value":1}, {"field":"name","operator" :"=","value":"jake"}]
        :param mod:  and,or
        :return: id>1 or name="jake"
        """
        whe = ""
        for i in range(len(wheres)):
            if "field" in wheres[i]:
                field = wheres[i]["field"]
            if "operator" in wheres[i]:
                operator = wheres[i]["operator"]
            else:
                operator = "="
            if "value" in wheres[i]:
                value = wheres[i]["value"]

            # 判断sql查询条件是否为str，是需要加上""查询
            if isinstance(value, str):
                value = "\"{}\"".format(value)

            whe = whe + field + operator + str(value)

            if i != len(wheres) - 1:
                whe = "{} {} ".format(whe, mod)
                # whe = whe + " " + mod + " "
        return whe

    # 通过where查询sql表中的数据
    def select_where_all(self,wheres=False, table_title=False, table_name=False, database_name=False):
        """
        sql拼接where查询
        :param wheres: 查询条件，条件格式为：[{"field":"id","operator" :">","value":1}, {"field":"name","operator" :"=","value":"jake"}]
        :param table_title: 表的列名 相当于后面的*  select * from table_name
        :param table_name: 表名
        :param database_name: 数据库名
        :return: 返回sql查询到的结果
        """
        if database_name == False:
            database_name = self.database
            self.get_all_db("use {}".format(database_name))
        if table_name == False:
            try:
                table_name = self.login["table"]
            except:
                raise myERROR("请输入查询的数据表名table_name")

        if table_title == False:
            table_title = "*"
        else:
            table_title = self.key_str(table_title)

        if wheres and wheres != []:
            try:
                wheredata = self.where_str(wheres)
                # wheredata = self.key_str(whereList, 'and')
                sql = "select {} from {} where {}".format(table_title, table_name, wheredata)
            except:
                raise myERROR("请按照此方式输入：[{\"field\":\"key\",\"operator\":\"=\",\"value\":\"value\"}]")

        else:
            sql = "select {} from {} ".format(table_title, table_name)

        print("查询sql语句为：{}".format(sql))
        try:
            result = self.get_all_db(sql)
        except:
            result = []



               
        return result

    # 插入数据表
    def table_insert(self,datas,table_name = False):
        """
        插入数据表数据
        :param datas: [{"name": "小明", "age": 18, "tel": 13567823931}, {"name": "小张", "age": '', "tel": 13567823931}]
        :param table_name: 表名  传：就是传的表名   不传：默认配置文件中的表名
        :return: 输出成功或者失败的sql语句
        """
        erro_sql = []
        ok_sql = []
        if table_name:
            table_values = self.tabale_name()
            if table_name not in table_values:
                raise myERROR("{}表不在当前数据库中".format(table_name))
        else:
            table_name = self.login["table"]
        for values in datas:
            sql_key = []
            sql_value = []

            for key, value in values.items():
                sql_key.append(key)
                sql_value.append(value)
                sql_key_data = self.key_str(sql_key)
                sql_value_data = self.value_str(sql_value)
            sql = "insert into {}({}) values({})".format(table_name, sql_key_data, sql_value_data)
            try:
                self.run_db(sql)
                ok_sql.append(sql)
            except:
                erro_sql.append(sql)
        if erro_sql:
            return "插入的sql有误：{}".format(erro_sql)
        else:
            return "成功的sql：{}".format(ok_sql)

    # 新建数据库
    def database_create(self,data=False):
        """
        新建数据库
        :param data: True:新建数据库语句   False:取yaml文件配置的新建数据库sql语句
        :return: 无
        """
        if data == False:
            markdir_datas= yaml_run.yaml_read_sql_database_makdir
            switch = markdir_datas["switch"]
            data = markdir_datas[switch]
        if isinstance(data,list):
            datas = data
        else:
            datas = []
            datas.append(data)
        for values in datas:
            self.run_db(values)

    # 新建数据表
    def table_create(self, data=False ,database_name = False):
        """
        新建数据表
        :param data: True:新建数据表语句   False:取yaml文件配置的新建数据表sql语句
        :param database_name: True ：选择的数据库新建表   False：当前登录的数据库新建表
        :return: 无
        """
        if data == False:
            markdir_datas= yaml_run.yaml_read_sql_table_makdir
            switch = markdir_datas["switch"]
            data = markdir_datas[switch]
            print(data)

        if database_name and database_name in self.database_name():
            self.database_use(database_name)

        if isinstance(data, list):
            datas = data
        else:
            datas = []
            datas.append(data)
        for values in datas:
            self.run_db(values)

    # 关闭数据库连接
    def close(self):
        # 关闭游标
        self.cur.close()
        # 关闭数据库连接
        self.conn.close()

if __name__ == '__main__':
    # # myDb = MyDb("goto")
    mydb = MyDb()
    # sql = 'select * from oks'
    # # sql = 'use tear'
    # # result = mydb.get_count_db(sql)
    # result = mydb.tabale_desc()
    # result = mydb.tabale_title(result)
    # result = mydb.key_str(["id","user"])
    # result = mydb.select_all()
    # result = mydb.where_str([{"field":"id","operator" :">","value":1}, {"field":"name","operator" :"=","value":"jake"}])
    # result = mydb.select_where_all([{"field":"id","operator" :"=","value":1}, {"field":"name","operator" :"=","value":"jake"}])
    # sql = 'insert into oks(name,age,tel) values("小明",Null,"13567823931")'
    # sql = ["小明", '', 13567823931]
    # result = mydb.value_str(sql)
    # result = mydb.runDb()
    # sql = {"name": "小明", "age": 18, "tel": 13567823931}
    # result = mydb.value_sql_str(sql)
    # print(mydb.select_all(["count(id) as 总数"]))
    # sql = [{"name": "小明", "age": 18, "tel": 13567823931}, {"name": "小张", "age": '', "tel": 13567823931}]
    # result = mydb.table_insert(sql)
    # print(mydb.select_all(["count(id) as 总数"]))
    # sql = yaml_run.yaml_read_sql_database_makdir["database_1"]
    # mydb.database_create(sql)
    # mydb.table_create(database_name = 'missjack')

    # print(result)
    #

    # result = mydb.run_db(sql)
    # result = mydb.database_name()
    # print(result)
    # print(mydb.database_name())
    # print(mydb.table_create())

    mydb. close()

    print(yaml_run.yaml_read_data)






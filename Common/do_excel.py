# -*- coding: utf-8 -*-
# @Author: Administrator
# @Date:   2022-03-13 21:02:20
# @Last Modified by:   Administrator
# @Last Modified time: 2022-04-10 20:30:16
"""
DoExcel  封装excel类
read_excel  读取工作表的单元格数据
write_excel  将值写入单元格
all_excel	读取工作表的所有单元格数据
max_row    读取最大行
clos_excle	读取最大列
sava_data	保存工作表
"""

import os
import json
from openpyxl import load_workbook,Workbook

from Common import do_path
from Common.do_path import data_path
# from Common.my_db import MyDb

class DoExcel:
	def __init__(self,paths,sheet,flag = False):
		"""
		excel处理类
		:param paths: flag：False  excel完成的工作路径     flag：存放在TestDatas目录下True excel的文件名
		:param sheet: 工作表名称
		:param flag: 开关  主要控制paths路径
		"""

		if flag:
			paths = os.path.join(do_path.data_path,paths)
		self.paths = paths
		self.sheet = sheet
		self.wb = load_workbook(self.paths)
		self.sheet = self.wb[self.sheet]

	def read_excel(self,row,column):
		"""
		read_excel:读取单元格的值
		row:单元格行数
		column:单元格列
		"""

		values = self.sheet.cell(row,column).value
		return values

	def write_excel(self,row,column,values):
		"""
		write_excel:修改单元格的值
		row:单元格行数
		column:单元格列
		values:需要修改的值
		"""

		self.sheet.cell(row,column).value = values
		self.wb.save(self.paths)

	def title_write(self, titleNames, rows=1):
		"""
		新建Excel标签行
		:param titleNames: 标题列名
		:param rows: 标签行
		:return:
		"""
		if isinstance(titleNames, list):
			titles = titleNames
		else:
			titles = []
			titles.append(titleNames)
		for i in range(len(titleNames)):
			self.write_excel(rows, i + 1, titles[i])

		self.wb.save(self.paths)

	def all_excel(self):  #推荐使用
		# 定义一个空列表,获取title标签名
		lists = []
		for item in list(self.sheet.rows)[0]:
			lists.append(item.value)

		# 放已经转换好的测试数据
		all_list = []
		for item in list(self.sheet.rows)[1:]:
			# 定义一个空列表放测试数据
			list_data = []
			for i in item:
				list_data.append(i.value)
			# dict(zip(value1,value2))  zip将value1和valu2组合，dict将zip组合的数据以字典的形式呈现
			res = dict(zip(lists,list_data))

			# json.loads 将字符串转为json格式
			# res["request_data"] = json.loads(res["request_data"])
			# res["expected"] = json.loads(res["expected"])
			# json.dumps 将json格式转为字符串
			# res["request_data"] = json.dumps(res["request_data"])

			all_list.append(res)
		return(all_list)


	def all_excel_one(self):
		# 定义一个空列表,获取title标签名
		lists = []
		for item in list(self.sheet.rows)[0]:
			lists.append(item.value)

		# 定义空列表放所有字典数据
		all_list = []
		# 取除了标题行的所有行的数据
		for item in list(self.sheet.rows)[1:]:
			dicts = {}
			# 对每一行进行循环得到每个单元格的值
			for index in range(len(item)):
				dicts[lists[index]] = item[index].value
			# excel读取的都是字符串或者数据,利用eval去掉字符串得到原有格式
			# dicts["msg"] = eval(dicts["msg"])
			# dicts["request_data"] = json.loads(dicts["request_data"])
			# dicts["expected"] = json.loads(dicts["expected"])
			all_list.append(dicts)

		return(all_list)

	def all_excel_two(self):
		# 最大行
		row = self.sheet.max_row

		# 最大列
		column = self.sheet.max_column
		lists = []
		for item in range(column):
			lists.append(self.sheet.cell(1,item+1).value)
		all_lists = []
		for item in range(2,row+1):
			dicts = {}
			for i in range(1,column+1):
				dicts[lists[i-1]] = self.sheet.cell(item,i).value
			# dicts["msg"] = eval(dicts["msg"])
			# dicts["request_data"] = json.loads(dicts["request_data"])
			# dicts["expected"] = json.loads(dicts["expected"])
			all_lists.append(dicts)

		return(all_lists)

	def max_row(self):
		# 最大行
		row = self.sheet.max_row
		return row


	def max_column(self):
		# 最大列
		column = self.sheet.max_column
		return column

	def clos_excle(self):
		self.wb.close()

	def sava_data(self):
		self.wb.save(self.paths)


class CreateExcel:
	def create_excel(self,paths,flag = False):
		"""
		创建excel工作簿
		:param paths: 工作簿创建位置，1.如果只填工作簿名称，默认创建在TestDatas目录下  2.如果是绝对路径就创建给定路径下
		:param flag: 去重开关，True 创建如果存在直接覆盖     Flase 会判断存在，存在不进行创建，不存在进行创建
		:return:执行操作
		"""
		wb = Workbook()
		if os.path.isfile(paths):
			file_path = paths
		else:
			file_path = os.path.join(do_path.data_path,paths)

		if flag:
			wb.save(file_path)
		else:
			if os.path.exists(file_path):
				print("文件{}的excle已经存在".format(file_path))
			else:
				wb.save(file_path)

	def create_sheet(self,paths,sheet_name,index_flag = False):
		"""
		创建工作表，如果存在不会进行创建
		:param paths: 需要创建工作表的工作簿地址，如果只传工作簿名称，默认去TestDatas目录查找
		:param sheet_name: 工作表名称
		:param index_flag: 工作表索引位置，False 默认末尾插入
		:return: 执行操作
		"""
		if os.path.isfile(paths):
			file_path = paths
		else:
			file_path = os.path.join(do_path.data_path,paths)

		wb = load_workbook(file_path)
		sheet_all = wb.sheetnames
		if sheet_name in sheet_all:
			print("工作表{}已经存在".format(sheet_name))
		else:
			if index_flag:
				index = index_flag
			else:
				index = len(sheet_all) + 1
			wb.create_sheet(index=index, title=sheet_name)
			wb.save(file_path)

	def copy_sheet(self,paths,sheet_name,org_sheet_name = False):
		"""
		复制工作表
		:param paths: 工作簿位置，1.如果只填工作簿名称，默认查找TestDatas目录  2.如果是绝对路径就创建给定路径下
		:param sheet_name: 工作表名称
		:param org_sheet_name: 默认 复制的工作表名称 + Copy  为True时，复制后工作表名称
		:return:执行操作
		"""
		if os.path.isfile(paths):
			file_path = paths
		else:
			file_path = os.path.join(do_path.data_path,paths)

		wb = load_workbook(file_path)
		sheet_all = wb.sheetnames
		if sheet_name in sheet_all:
			cp_sheet = wb.copy_worksheet(wb[sheet_name])
			if org_sheet_name:
				cp_sheet.title = org_sheet_name
			print(cp_sheet)
			print(wb[sheet_name])
		wb.save(file_path)

	def rename_sheet(self,paths,sheet_name,org_sheet_name):
		"""
		重命名工作表
		:param paths: 工作簿位置，1.如果只填工作簿名称，默认查找TestDatas目录  2.如果是绝对路径就创建给定路径下
		:param sheet_name: 重命名前工作表名称
		:param org_sheet_name: 重命名后工作表名称
		:return:执行操作
		"""
		if os.path.isfile(paths):
			file_path = paths
		else:
			file_path = os.path.join(do_path.data_path,paths)

		wb = load_workbook(file_path)
		sheet_all = wb.sheetnames
		if sheet_name in sheet_all:
			wb[sheet_name].title = org_sheet_name
		wb.save(file_path)

	def all_sheet(self,paths):
		"""
		拉取工作簿下的所有工作表
		:param paths: 工作簿位置，1.如果只填工作簿名称，默认查找TestDatas目录  2.如果是绝对路径就创建给定路径下
		:return:列表返回工作簿下的所有工作表
		"""
		if os.path.isfile(paths):
			file_path = paths
		else:
			file_path = os.path.join(do_path.data_path,paths)

		wb = load_workbook(file_path)
		sheet_all = wb.sheetnames
		return sheet_all

	def del_sheet(self,paths,sheet_name):
		"""
		删除工作表
		:param paths: 工作簿位置，1.如果只填工作簿名称，默认查找TestDatas目录  2.如果是绝对路径就创建给定路径下
		:param sheet_name: 删除的工作表名称
		:return:执行操作
		"""
		if os.path.isfile(paths):
			file_path = paths
		else:
			file_path = os.path.join(do_path.data_path,paths)

		wb = load_workbook(file_path)
		sheet_all = wb.sheetnames
		if sheet_name in sheet_all:
			del wb[sheet_name]
		wb.save(file_path)

	def del_excel(self,paths):
		if os.path.isfile(paths):
			file_path = paths
			if os.path.exists(file_path):
				os.remove(file_path)
			else:
				print("删除的{}文件不存在".format(file_path))
		else:
			file_path = os.path.join(do_path.data_path,paths)
			if os.path.isfile(file_path) and os.path.exists(file_path):
				os.remove(file_path)
			else:
				print("删除的{}文件不存在".format(file_path))

if __name__ == '__main__':

	# file_path = do_path.data_path + r"\video.xlsx"
	# print(file_path)
	# do_excel = DoExcel(file_path,"视频号测试数据20221001-20221017")


	# # 读取1,1单元格的值
	# values = do_excel.read_excel(1,1)
	# print(values)

	# # 写入1,1单元格的值
	# do_excel.write_excel(1,1,"username")
	
	# # 修改后，再读取1,1单元格的值
	# values = do_excel.read_excel(1,1)
	# print(values)

	# values = do_excel.all_excel()

	# values = do_excel.all_excel()
	# print(values)

	# 关闭excel
	# do_excel.clos_excle()

	# paths = r"D:\chen\python\ckt\temporary\abc.xlsx"
	paths = r"D:\chen\python\ckt\temporary\ok\bc"
	# paths = r"acbc.xlsx"
	# mod = "Sheet21"
	cr_excle = CreateExcel()
	# cr_excle.create_excel(paths)
	# cr_excle.create_sheet(paths,mod)
	# cr_excle.del_excel(paths)

	# cr_excle.copy_sheet(paths,"Sheet","ok")
	# cr_excle.rename_sheet(paths,"ok","Sheet")

	# all_sheet = cr_excle.all_sheet(paths)
	# print(all_sheet)
	# cr_excle.del_sheet(paths,all_sheet[len(all_sheet)-1])
	# all_sheet = cr_excle.all_sheet(paths)
	# print(all_sheet)





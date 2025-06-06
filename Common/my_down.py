# -*- coding: utf-8 -*-
# @Author: v_ghuachen
# @Date:   2022-01-11 14:22:41
# @Last Modified by:   v_ghuachen
# @Last Modified time: 2022-02-16 15:35:14

"""
1.download_file  通过url下载视频类
2.根据视频截取帧数
"""

# you-get -o d:/media http://adsmind.gdtimg.com/ads_svp_video__0b53gaabsaaap4aizuo2w5qrumaedeyaagka.f0.mp4?dis_k=2e948464e8ffafcf9d07bc357c6e0f1c&dis_t=1641715024&m=a9b24b38c88c0e83d923b8ab9c1ee22a

# -*- coding: utf-8 -*-


# 下载

import requests

from contextlib import closing

import time
import cv2
import re

from openpyxl import load_workbook

import os

from Common.do_path import data_path
from Common.do_excel import DoExcel


# print(sheet.max_row)
# print(sheet.max_column)

def download_file(url, paths ,flag = "ok"):
    """
    :param url: 下载素材地址
    :param paths: 保存文件路径（精确到文件名）
    :param flag: 决定什么文件后缀
    flag 不填或者ok 按照链接文件下载
    flag  填了其他格式，按照其他格式进行下载
    flag  为false 按照paths传参路径进行下载
    :return:
    """
    with closing(requests.get(url, stream=True)) as r:
        chunk_size = 1024
        # 获取素材文件信息
        head = r.headers
        # print(head)
        # print(url)
        # 取出文件大小
        content_size = int(r.headers['content-length'])
        format = r.headers['Content-Type']
        file_type = str.split(format,'/')[1]
        if flag == "ok":
            path = "{}.{}".format(paths,file_type)
        elif flag:
            path = "{}.{}".format(paths,flag)
        else:
            path = paths

        print('下载开始')
        with open(path,"wb") as f:
            n = 1
            for chunk in r.iter_content(chunk_size=chunk_size):
                loaded = n * 1024.0 / content_size
                f.write(chunk)
                # print('已下载{0:%}'.format(loaded))
                n += 1
            print('下载完成{0:%}'.format(loaded))

def extract_video(filename,dst,flag = 1000):
    """
    根据视频截取帧数
    :param filename: 视频文件路径
    :param dst: 存储截取图片存储路径，默认格式jpg，文件路径需要带上文件名
    :param flag: 截帧频率，默认按照1000截取一帧
    :return:
    """
    # 按照1000帧截取1张
    interval = int(flag)
    # c 起始位置
    c = 1
    # 按照多少帧率累加
    num = 1
    # 读取视频文件
    vc = cv2.VideoCapture(filename)
    if vc.isOpened():
        # 按帧读取
        # val,frame是获vc.read()方法的两个返回值，其中val是布尔值，如果读取帧是正确的则返回True，如果文件读取到结尾，它的返回值就为False。frame就是每一帧的图像，是个三维矩阵
        val,frame = vc.read()
    else:
        val = False

    while val:
        val,frame = vc.read()
        if c % interval == 0:
            # print(str(c),"1")
            # 存储截取矩阵
            cv2.imwrite(dst + str(c) + '.jpg',frame)
        c = c + num
        # print(c)
        # 持续时间，展示1帧
        cv2.waitKey(1)

    # 释放资源
    vc.release()




if __name__ == "__main__":
    # url = "http://media.cugbonline.cn/dest/cd7/cd7f62da-c976-4163-8494-630e430bb45a.mp4"
    # download_file(url, "D:\\media/abs/a"+str(int(time.time()))+".mp4")

    # path = os.path.dirname(os.getcwd()) +r"\TestDatas\red.xlsx"
    # # print(path)
    # down_path = r"D:\media\前测二期\红灯\红灯_"
    #
    #
    # wb = load_workbook(path)
    # sheet = wb['media']
    # times = sheet.max_row
    # # print(times)
    # # print(times)
    # for k in range(1,times):
    #     values = sheet.cell(k + 1, 4).value
    #     width = sheet.cell(k+1,3).value
    #     height = sheet.cell(k+1,2).value
    #     # print(values)
    #     download_file(str(values), down_path + str(width) + "_" + str(height) +"_"+ str(int(time.time()))+".mp4")


    path  = os.path.join(data_path,"shals.xlsx")
    do_my = DoExcel(path,"Sheet1")
    value = do_my.all_excel()

    urlpath = r"D:\media\黑产\黑产updata"
    for i in range(len(value)):
        md5 = value[i]["md5"]
        name = value[i]["name"]
        url = value[i]["url"]
        orpath = os.path.join(urlpath,md5)
        if os.path.exists(orpath) == False:
            os.mkdir(orpath)

        filepath = os.path.join(orpath,name)
        download_file(url,filepath)


import base64
import os
import hashlib
import time
import requests
import random
import threading
from contextlib import closing
from openpyxl import load_workbook,Workbook
import cv2
import shutil
import yaml

# from Common.do_path import data_path


class DoExcel:
    def __init__(self, paths, sheet, flag=False):
        """
        excel处理类
        :param paths: flag：False  excel完成的工作路径     flag：存放在TestDatas目录下True excel的文件名
        :param sheet: 工作表名称
        :param flag: 开关  主要控制paths路径
        """

        if flag:
            paths = os.path.join(data_path, paths)
        self.paths = paths
        self.sheet = sheet
        self.wb = load_workbook(self.paths)
        self.sheet = self.wb[self.sheet]

    def read_excel(self, row, column):
        """
        read_excel:读取单元格的值
        row:单元格行数
        column:单元格列
        """

        values = self.sheet.cell(row, column).value
        return values

    def write_excel(self, row, column, values):
        """
        write_excel:修改单元格的值
        row:单元格行数
        column:单元格列
        values:需要修改的值
        """

        self.sheet.cell(row, column).value = values
        self.wb.save(self.paths)

    def title_write(self, titleNames, rows=1):
        """
        新建Excel标签行
        :param titleNames: 标题列名
        :param rows: 标签行
        :return:
        """
        if isinstance(titleNames, list):
            titles = titleNames
        else:
            titles = []
            titles.append(titleNames)
        for i in range(len(titleNames)):
            self.write_excel(rows, i + 1, titles[i])

        self.wb.save(self.paths)

    def all_excel(self):  # 推荐使用
        # 定义一个空列表,获取title标签名
        lists = []
        for item in list(self.sheet.rows)[0]:
            lists.append(item.value)

        # 放已经转换好的测试数据
        all_list = []
        for item in list(self.sheet.rows)[1:]:
            # 定义一个空列表放测试数据
            list_data = []
            for i in item:
                list_data.append(i.value)
            # dict(zip(value1,value2))  zip将value1和valu2组合，dict将zip组合的数据以字典的形式呈现
            res = dict(zip(lists, list_data))

            # json.loads 将字符串转为json格式
            # res["request_data"] = json.loads(res["request_data"])
            # res["expected"] = json.loads(res["expected"])
            # json.dumps 将json格式转为字符串
            # res["request_data"] = json.dumps(res["request_data"])

            all_list.append(res)
        return (all_list)

    def all_excel_one(self):
        # 定义一个空列表,获取title标签名
        lists = []
        for item in list(self.sheet.rows)[0]:
            lists.append(item.value)

        # 定义空列表放所有字典数据
        all_list = []
        # 取除了标题行的所有行的数据
        for item in list(self.sheet.rows)[1:]:
            dicts = {}
            # 对每一行进行循环得到每个单元格的值
            for index in range(len(item)):
                dicts[lists[index]] = item[index].value
            # excel读取的都是字符串或者数据,利用eval去掉字符串得到原有格式
            # dicts["msg"] = eval(dicts["msg"])
            # dicts["request_data"] = json.loads(dicts["request_data"])
            # dicts["expected"] = json.loads(dicts["expected"])
            all_list.append(dicts)

        return (all_list)

    def all_excel_two(self):
        # 最大行
        row = self.sheet.max_row

        # 最大列
        column = self.sheet.max_column
        lists = []
        for item in range(column):
            lists.append(self.sheet.cell(1, item + 1).value)
        all_lists = []
        for item in range(2, row + 1):
            dicts = {}
            for i in range(1, column + 1):
                dicts[lists[i - 1]] = self.sheet.cell(item, i).value
            # dicts["msg"] = eval(dicts["msg"])
            # dicts["request_data"] = json.loads(dicts["request_data"])
            # dicts["expected"] = json.loads(dicts["expected"])
            all_lists.append(dicts)

        return (all_lists)

    def max_row(self):
        # 最大行
        row = self.sheet.max_row
        return row

    def max_column(self):
        # 最大列
        column = self.sheet.max_column
        return column

    def clos_excle(self):
        self.wb.close()

    def sava_data(self):
        self.wb.save(self.paths)


class MyOs:
    def filelist(self, file_dir):
        """
        查找文件夹下的所有文件，包含子目录、孙子目录下所有的文件
        :param file_dir: 文件夹路径
        :return: 返回文件列表
        """
        if os.path.exists(file_dir) == False:
            # 判断文件路径是否存在，不存在直接返回
            return "{}文件路径".format(file_dir)
        # 拉出根目录下的所有文件
        path_list = []
        for root, dirs, files in os.walk(file_dir):
            # root 根目录
            # dirs 子目录
            # files 子目录文件
            # 主要拉出根路径的文件
            if root == file_dir and dirs != []:
                # print(root)
                for i in range(len(files)):
                    paths = os.path.join(root, files[i])
                    path_list.append(paths)
                    # print(paths)
                # print("\n")
            # 主要拉出根目录下的所有子目录的文件
            if dirs == []:
                # print('目录路径    ：',root)
                # print('子目录名称  ：',dirs)
                # print('目录下的文件：',files)
                # print("\n")
                # print(root)
                for i in range(len(files)):
                    paths = os.path.join(root, files[i])
                    path_list.append(paths)
                    # print(paths)
                # print("\n")
        return path_list

    def findStype(self, file_dir, stype="txt"):
        """
        查找文件后缀为指定格式文件
        :param file_dir: 文件路径
        :param stype: 文件格式 默认 txt
        :return: 输出符合条件的路径列表
        """
        if isinstance(file_dir, list):
            # 判断是否为列表
            file_dirs = file_dir
        elif isinstance(file_dir, str) and os.path.exists(file_dir):
            # 判断是否为字符串，并且文件路径是否存在
            file_dirs = []
            file_dirs.append(file_dir)
        else:
            print("输入的文件路径{}不对".format(file_dir))
            return []

        values = []
        for files in file_dirs:
            if os.path.exists(files):
                filesList = files.split(".")
                if filesList[len(filesList) - 1] == stype:
                    values.append(files)
        return values

    def outfile(self, fileFront):
        """
        随机更改文件，文件，字符
        :param moveFront: 需要更改的文件、文件夹路径
        :return:
        """
        outstr = ''
        if os.path.isfile(fileFront):
            # 文件随机命名
            value = fileFront.split('.')
            outstr = "{}{}_副本{}{}{}".format(outstr, value[0], random.randint(1, 9), random.randint(0, 9),
                                            random.randint(0, 9))
            for i in range(1, len(value) - 1):
                outstr = outstr + '.' + value[i]
            outstr = "{}.{}".format(outstr, value[len(value) - 1])

        elif os.path.isdir(fileFront):
            # 文件夹随机命名
            value = fileFront.split('/')
            for i in range(0, len(value) - 1):
                outstr = outstr + value[i] + '/'
            outstr = "{}{}_副本{}{}{}".format(outstr, value[len(value) - 1], random.randint(1, 9), random.randint(0, 9),
                                            random.randint(0, 9))

        else:
            # 字符随机
            outstr = "{}_副本{}{}{}".format(fileFront, random.randint(1, 9), random.randint(0, 9), random.randint(0, 9))

        return outstr

    def movefile(self, moveFront, moveAfter):
        """
        移动文件、文件夹
        :param moveFront: 需要移动的文件、文件夹路径
        :param moveAfter: 移动后的文件、文件夹路径
        :return:
        """
        try:
            shutil.move(moveFront, moveAfter)
        except:
            print("{} 在文件 {} 已存在".format(moveFront, moveAfter))
            return moveFront

    def copyfile(self, copyFront, copyAfter, overlay=True):
        """
        复制文件、文件夹
        :param copyFront: 需要复制的文件、文件夹路径
        :param copyAfter: 复制后的文件、文件夹路径
        :param overlay: True 复制    False 覆盖复制
        :return: 无
        """
        if overlay:
            shutil.copy(copyFront, copyAfter)
        else:
            shutil.copy2(copyFront, copyAfter)

    def renamefile(self, nameFront, nameAfter):
        """
        文件、文件夹重命名
        :param nameFront: 需要命名的文件、文件夹路径
        :param nameAfter: 命名后的文件、文件夹路径
        :return: 无
        """
        if os.path.isdir(nameFront) and os.path.isdir(nameAfter):
            os.rename(nameFront, nameAfter)
        elif os.path.isfile(nameFront) and os.path.isfile(nameAfter):
            os.rename(nameFront, nameAfter)
        else:
            print("{} 和 {}不能进行重命名".format(nameFront, nameAfter))

    def readfile(self, file_path):
        """
        读取文件文件内容
        :param file_path: 文本文件路径
        :return: 返回文本文件内容
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            for line in lines:
                return line

        f.close()

    def writefile(self, file_path, txt, isw=False):
        """
        写文本文件
        :param file_path: 问你路径
        :param txt: 写入的文本
        :param isw: True 覆盖写入   False 追加写入
        :return: 无
        """
        if isw:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(txt)
        else:
            with open(file_path, 'a', encoding='utf-8') as f:
                f.write("\n" + txt)
        f.close()

class MyCsv:
    def __init__(self,path,name,langue = 'utf-8',flag = False):
        self.path = path
        self.name = name
        self.paths = os.path.join(path,name)
        self.langue = langue


        # 判断文件名称是否存在文件目录下，不存在则新建
        plist = os.listdir(path)
        if str(name) not in plist:
            self.title("a","b")

    def readCsv(self,swith_to = "#",default = "None",mode ='r'):
        with open(self.paths,mode = mode,encoding= self.langue) as f:
            txtlines = f.readlines()
        title = txtlines[0].strip().split(',')

        all_list = []
        title_num = len(title)

        for i in range(1,len(txtlines)):
            # strip 截取首尾空字符
            temp = txtlines[i].strip()
            # startswith查找什么开头
            if temp.startswith(swith_to):
                continue
            # 以,号为分隔符截取字段
            temp_list = temp.split(',')
            num = len(temp_list)
            dict_list = {}
            for j in range(title_num):
                if j < int(num):
                    dict_list[title[j]] = temp_list[j]
                else:
                    dict_list[title[j]] = default

            all_list.append(dict_list)


        return all_list

    def readAll(self,mode = 'r'):
        with open(self.paths,mode = mode,encoding = self.langue) as f:
            value = f.read()
            return value


    def dowrite(self,*args):
        txt = ''
        for i in range(len(args)):
            if i == 0:
                txt = "{}".format(args[i])
            else:
                txt = "{},{}".format(txt,args[i])

        with open(self.paths,mode = 'a',encoding = self.langue) as f:
            f.write('\n{}'.format(txt))

    def title(self,*args):
        txt = ''
        for i in range(len(args)):
            # 查找字符重复数
            counts = txt.count(args[i])
            # if counts > 0:
            #     raise Exception("{}字段重复".format(args[i]))
            if i == 0:
                txt = "{}".format(args[i])
            else:
                txt = "{},{}".format(txt,args[i])
        with open(self.paths,mode = 'w',encoding = self.langue) as f:
            f.write('{}'.format(txt))


class MyYaml:
    def __init__(self, path, name, langue='utf-8', flag=False):
        self.path = path
        self.name = name
        self.paths = os.path.join(path, name)
        self.langue = langue

        # print(self.paths)

        # 判断文件名称是否存在文件目录下，不存在则新建
        plist = os.listdir(path)
        if str(name) not in plist:
            with open(self.paths, mode='w', encoding=self.langue) as f:
                f.write('')

    def readYaml(self):
        with open(self.paths, mode='r', encoding=self.langue) as f:
            conf = yaml.load(f, Loader=yaml.FullLoader)
        return (conf)

    def writeYaml(self, datas, mode='w'):
        with open(self.paths, mode=mode, encoding=self.langue) as f:
            conf = yaml.dump(datas, allow_unicode=True)
            f.write(conf)


class Downs:
    def download_file(self, url, paths, flag="ok", ranflag='md5'):
        """
        url素材下载
        :param url: 下载素材地址
        :param paths: 保存文件路径（精确到文件名）
        :param flag: 决定什么文件后缀
        flag 不填或者ok 按照链接文件下载
        flag  填了其他格式，按照其他格式进行下载
        flag  为false 按照paths传参路径进行下载
        ranflag: md5：根据md5进行命名，可以去重    true：随机命名，不会去重  false：不进行随机命名
        :return:
        """
        try:
            # 获取url的宽高信息，可能会报错，所以用try
            width_hight = self.get_video_dimensions(url)
        except:
            # 如果报错，设置宽高为0
            width_hight = (0, 0)

        # 如果为md5，就用md5进行文件名称拼接，可以进行文件名去重
        if ranflag == 'md5':
            try:
                # 获取url的md5
                md5 = self.get_url_md5(url)
                # 拼接文件名称
                paths = "{}_{}_{}_{}".format(paths, width_hight[0], width_hight[1], md5)
            except:
                # 如果通过url获取不到md5，就不用md5拼接
                paths = "{}_{}_{}".format(paths, width_hight[0], width_hight[1])
                paths = self.name_rand(paths)
        # 如果开关不为md5，为true，文件名称就随机生成后缀
        elif ranflag:
            paths = "{}_{}_{}".format(paths, width_hight[0], width_hight[1])
            paths = self.name_rand(paths)
        else:
            pass

        with closing(requests.get(url, stream=True)) as r:
            chunk_size = 1024
            # 获取素材文件信息
            head = r.headers
            # print(head)
            # print(url)
            # 取出文件大小
            content_size = int(r.headers['content-length'])
            format = r.headers['Content-Type']
            file_type = str.split(format, '/')[1]
            if flag == "ok":
                path = "{}.{}".format(paths, file_type)
            elif flag:
                path = "{}.{}".format(paths, flag)
            else:
                path = paths

            print('下载开始')
            with open(path, "wb") as f:
                n = 1
                for chunk in r.iter_content(chunk_size=chunk_size):
                    loaded = n * 1024.0 / content_size
                    f.write(chunk)
                    # print('已下载{0:%}'.format(loaded))
                    n += 1
                print('下载完成{0:%}'.format(loaded))

    def download_file_mkdir(self, url, paths, flag="ok", ranflag="md5"):
        """
        根据url宽高创建对应素材进行下载到对应尺寸的文件夹中
        :param url: 下载素材地址
        :param paths: 保存文件路径（精确到文件名）
        :param flag: 决定什么文件后缀
        flag 不填或者ok 按照链接文件下载
        flag  填了其他格式，按照其他格式进行下载
        flag  为false 按照paths传参路径进行下载
        ranflag: md5：根据md5进行命名，可以去重    true：随机命名，不会去重  false：不进行随机命名
        :return:
        """
        split_symbol = '\\'
        try:
            # 通过spile将文件路径中的文件名提取出来
            split_value = self.str_split(paths,split_symbol)
            # 获取url的宽高信息
            width_hight = self.get_video_dimensions(url)
            # 返回文件路径的上一级文件夹目录
            out_start_path = os.path.dirname(paths)
            # 拼接以“宽——高”的文件目录名称
            out_center_path = "{}_{}".format(width_hight[0],width_hight[1])
            # print(split_value)
            # spile提取文件名是否为空
            if split_value != []:
                # 取出文件名
                out_end_path = split_value[-1]
                # 拼接新的文件夹目录，增加以“宽——高”的文件目录
                out_path = os.path.join(out_start_path, out_center_path)
                # 调用is_path函数，文件夹路径不存在就创建
                self.is_path(out_path)

                # 拼接新的文件目录
                paths = os.path.join(out_path,out_end_path)
            else:
                # 提示未找到
                print("路径{}未找到为{}的分隔符".format(paths,split_symbol))
        except:
            pass

        try:
            # 获取url的宽高信息，可能会报错，所以用try
            width_hight = self.get_video_dimensions(url)
        except:
            # 如果报错，设置宽高为0
            width_hight = (0,0)
        # 如果为md5，就用md5进行文件名称拼接，可以进行文件名去重
        if ranflag == 'md5':
            try:
                # 获取url的md5
                md5 = self.get_url_md5(url)
                # 拼接文件名称
                paths = "{}_{}_{}_{}".format(paths,width_hight[0],width_hight[1],md5)
            except:
                # 如果通过url获取不到md5，就不用md5拼接
                paths = "{}_{}_{}".format(paths, width_hight[0], width_hight[1])
                paths = self.name_rand(paths)
        # 如果开关不为md5，为true，文件名称就随机生成后缀
        elif ranflag:
            paths = "{}_{}_{}".format(paths, width_hight[0], width_hight[1])
            paths = self.name_rand(paths)
        else:
            pass

        with closing(requests.get(url, stream=True)) as r:
            chunk_size = 1024
            # 获取素材文件信息
            head = r.headers
            # print(head)
            # print(url)
            # 取出文件大小
            content_size = int(r.headers['content-length'])
            format = r.headers['Content-Type']
            file_type = str.split(format, '/')[1]
            if flag == "ok":
                path = "{}.{}".format(paths, file_type)
            elif flag:
                path = "{}.{}".format(paths, flag)
            else:
                path = paths

            print('下载开始')
            with open(path, "wb") as f:
                n = 1
                for chunk in r.iter_content(chunk_size=chunk_size):
                    loaded = n * 1024.0 / content_size
                    f.write(chunk)
                    # print('已下载{0:%}'.format(loaded))
                    n += 1
                print('下载完成{0:%}'.format(loaded))

    def multi_thread(self, obj, data, num, flag=False):
        """
        并发的封装
        :param obj: 并发执行的函数
        :param data: 并非需要的数据
        :param num: 并发数
        :param flag: 用于执行obj函数传两个参数使用，默认一个参数
        :return: 执行
        """
        threads_list = []
        # flag :True 允许传两个参数
        if flag:
            for value in data:
                threads_list.append(threading.Thread(target=obj, args=(value, flag), name=f"进程{int(time.time())}"))
        else:
            for value in data:
                threads_list.append(threading.Thread(target=obj, args=(value,), name=f"进程{int(time.time())}"))
        # 开始线程计数
        num_one = 0
        # 结束线程计数
        num_two = 0
        # 每次执行的并发数
        num = int(num)
        # 开始线程计数，小于数据条数，就启动线程处理数据
        while num_one < len(data):
            # 每次启动线程的并发数
            for i in range(num):
                print("{}{}".format(threading.current_thread().name, num_one))
                threads_list[num_one].start()
                num_one += 1
                # 开始线程计数，大于数据条数，跳出循环
                if num_one >= len(data):
                    break
            # 每次等待线程的并发数结束
            for i in range(num):
                threads_list[num_two].join()
                num_two += 1
                # 开始线程计数，大于数据条数，跳出循环
                if num_two >= len(data):
                    break

    def is_path(self, path):
        """
        判断目录是否存在，不存在就进行新建
        :param path: 文件夹路径
        :return: 执行
        """
        # 判断文件夹是否存在
        if os.path.isdir(path) == False:
            os.makedirs(path)

    def is_path_file(self, path):
        """
        判断文件目录是否存在，不存在返回False 存在返回True
        :param path: 文件路径
        :return: 返回 True 或者 False
        """
        # 判断文件夹是否存在
        if os.path.isfile(path):
            return True
        else:
            return False

    def chagemd5(self,file_path):
        # 改变文件的MD5
        # filename 文件路径
        myfile = open(file_path, 'a')
        myfile.write("####&&&&")
        myfile.close

    def filemd5(self,file_path, flag=False):
        """
        读取文件MD5
        :param filename: 路径
        :param flag: 开关 True 改变MD5  False 不改变MD5 只读取MD5
        :return:返回MD5
        """

        if flag:
            chagemd5(file_path)
        hasher = hashlib.md5()
        with open(file_path, "rb") as file:
            buf = file.read()
            while len(buf) > 0:
                hasher.update(buf)
                buf = file.read()
        print(hasher.hexdigest())
        return hasher.hexdigest()

    def str_split(self,data,ide = '\\'):
        """
        数据中找到idt，找到返回找到的数据，没有找到返回[]
        :param data:数据源
        :param ide:按照ide对数据源切割
        :return:切割后的列表
        """
        data = str(data)
        # 找不到就返回[]
        if data.find(ide) == -1:
            return []
        else:
            value = data.split(ide)
            return value

    def name_rand(self, name="Resource", out_file_path=None, flag=True):
        """
        文件名称拼接随机值
        :param name: 文件名称
        :param out_file_path: True 进行文件路径拼接   Flase 不进行文件路径拼接
        :param flag: True 进行拼接  Flase 不进行拼接
        :return: 输出随机命名的文件或者路径
        """
        # 开关开生成随机文件名后缀
        if flag:
            # 生成文件名+随机后缀的组合文件名称
            file_name = "{}_{}{}{}{}".format(name, int(time.time()), random.randint(1, 9), random.randint(1, 9),
                                             random.randint(1, 9), random.randint(1, 9))
        else:
            file_name = name
        # 判断文件夹路径是否传递和文件夹是否存在，满足条件就进行文件目录+文件名称拼接路径
        if out_file_path != None and os.path.isdir(out_file_path):
            file_name = os.path.join(out_file_path, file_name)
        # 返回最新的路径
        return file_name

    def get_video_dimensions(self,url):
        # 通过视频ulr解析视频宽高
        video_capture = cv2.VideoCapture(url)
        width = int(video_capture.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(video_capture.get(cv2.CAP_PROP_FRAME_HEIGHT))
        video_capture.release()
        return width, height

    def get_url_md5(self,url):
        filename = url.split('/')[-1]  # 提取文件名

        response = requests.get(url)  # 发送GET请求

        md5_obj = hashlib.md5()  # 创建MD5对象
        md5_obj.update(response.content)  # 更新MD5对象

        return md5_obj.hexdigest()  # 返回MD5值

    def run_down_csv(self, input_file, csv_name, dict_name, out_file, num=3, file_name="Resource",flag_markdir = False):
        """
        csv文件批量下载
        :param input_file: 输入文件路径（不精确到文件名）
        :param csv_name: csv文件名
        :param dict_name: 数据源中需要下载的列名
        :param out_file: 输出文件路径
        :param num: 并发数
        :param file_name: 数据数据的名称
        :param flag_markdir True 执行download_file_mkdir函数，Flase 执行download_file函数
        :return: 执行下载命令
        """
        self.is_path(out_file)
        paths = self.name_rand(file_name, out_file, False)
        # print(input_file,csv_name)
        mycsv = MyCsv(input_file, csv_name, 'gbk')
        datas = mycsv.readCsv()
        url_datas = []
        for i in range(len(datas)):
            url_datas.append(datas[i][dict_name])
        if flag_markdir:
            self.multi_thread(self.download_file_mkdir, url_datas, num, paths)
        else:
            self.multi_thread(self.download_file, url_datas, num, paths)

    def run_down_excel(self, input_file, sheet_name, dict_name, out_file, num=3, file_name="Resource",flag_markdir = False):
        """
        csv文件批量下载
        :param input_file: 输入文件路径（精确到文件名）
        :param sheet_name: excel工作表名
        :param dict_name: 数据源中需要下载的列名
        :param out_file: 输出文件路径
        :param num: 并发数
        :param file_name: 数据数据的名称
        :param flag_markdir True 执行download_file_mkdir函数，Flase 执行download_file函数
        :return: 执行下载命令
        """
        self.is_path(out_file)
        paths = self.name_rand(file_name, out_file, 0)
        # print(input_file,sheet_name)
        myexcel = DoExcel(input_file, sheet_name)
        datas = myexcel.all_excel()
        # print(datas)
        url_datas = []
        for i in range(len(datas)):
            url_datas.append(datas[i][dict_name])

        if flag_markdir:
            self.multi_thread(self.download_file_mkdir, url_datas, num, paths)
        else:
            self.thread = self.multi_thread(self.download_file, url_datas, num, paths)

    def file_name_all(self,file_dir):
    # 拉出根目录下的所有文件
        path_list = []
        for root, dirs, files in os.walk(file_dir):
            # root 根目录
            # dirs 子目录
            # files 子目录文件
            # 主要拉出根路径的文件
            if root == file_dir and dirs != []:
                # print(root)
                for i in range(len(files)):
                    paths = os.path.join(root, files[i])
                    path_list.append(paths)
                    # print(paths)
                # print("\n")
            # 主要拉出根目录下的所有子目录的文件
            if dirs == []:
                # print('目录路径    ：',root)
                # print('子目录名称  ：',dirs)
                # print('目录下的文件：',files)
                # print("\n")
                # print(root)
                for i in range(len(files)):
                    paths = os.path.join(root, files[i])
                    path_list.append(paths)
                    # print(paths)
                # print("\n")
        return path_list






if __name__ == '__main__':
    file_path = "D:/chen/python/ckt/temporary/image"
    url = "https://pcm-30170.picsz.qpic.cn/product/20230714103519/b9ce7ad47f684e5055edc97070e42216.jpg"

    go_down = Downs()

    # kuan_heigh = go_down.get_video_dimensions(url)
    # print(kuan_heigh)

    # 单个url下载
    # go_down.download_file(url,file_path,ranflag=False)

    # csv文件下载
    go_down.run_down_csv('H:\腾讯\图片\房产','case.csv','FBaseImageUrl','H:\腾讯\图片\房产',3,"image",flag_markdir = True)

    # excel文件下载
    # go_down.run_down_excel('D:/chen/python/ckt/temporary/case.xlsx','Sheet4','url','D:/chen/python/ckt/temporary',3,"image")
